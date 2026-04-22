"""
Competitor Price Scraper — daily pricing across all matched competitors.

Strategy per competitor:
  - Sitemap discovery: fetch sitemap once, build title->URL map via slug fuzzy match
  - Direct product page scrape via Jina Reader (free, Cloudflare bypass)
  - Title fuzzy match >= TITLE_MATCH_THRESHOLD to confirm correct product page
  - Cost floor check: reject scraped price if > COST_FLOOR_TOLERANCE below our unit cost

Inputs:
  - CATALOG_PATH  : Price Vs Cost export (MPN + title + revenue + cost)
  - MATCHES_PATH  : Competitor Matching export (MPN -> competitor name)
  - CONFIG_PATH   : competitor_config.json built by competitor_discovery.py

Output:
  - Excel report to WORKAI/reports/ (synced to Google Drive automatically)

Usage:
  python price_scraper.py          # top 50 products (default)
  python price_scraper.py 500      # top 500 products
  python price_scraper.py 0        # all products (no cap)
"""

import html
import json
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime

import openpyxl
import requests
from thefuzz import fuzz

# ── Config ────────────────────────────────────────────────────────────────────

JINA_BASE             = "https://r.jina.ai/"
TITLE_MATCH_THRESHOLD = 72     # minimum fuzzy match % to accept a URL
COST_FLOOR_TOLERANCE  = 0.10   # reject if competitor price > 10% below our unit cost
TOP_N_PRODUCTS        = 50     # default scrape depth (override via CLI arg)

from config import CATALOG_PATH, MATCHES_PATH
CONFIG_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "competitor_config.json")
REPORTS_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports")


def _load_competitors() -> dict:
    """Load competitor config built by competitor_discovery.py."""
    if not os.path.exists(CONFIG_PATH):
        print(f"WARNING: {CONFIG_PATH} not found — run competitor_discovery.py first")
        return {}
    with open(CONFIG_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    out = {}
    for name, cfg in raw.items():
        if cfg.get("status") in ("ok", "low_count"):
            out[name] = {"domain": cfg["domain"], "sitemap": cfg.get("sitemap")}
        else:
            out[name] = {"domain": cfg.get("domain", ""), "sitemap": None}
    return out


COMPETITORS = _load_competitors()


# ── HTTP helpers ──────────────────────────────────────────────────────────────

def _fetch(url: str, timeout: int = 45) -> str | None:
    """Fetch a product page via Jina Reader (Cloudflare bypass, returns markdown)."""
    try:
        r = requests.get(
            JINA_BASE + url,
            headers={"Accept": "text/markdown", "X-Timeout": "30"},
            timeout=timeout,
        )
        if r.status_code == 200 and r.text:
            return r.content.decode("utf-8", errors="replace")
    except Exception as e:
        print(f"    [Jina] {e}")
    return None


def _fetch_direct(url: str, timeout: int = 15) -> str | None:
    """Direct HTTP fetch — used for sitemaps (no Jina needed)."""
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return None


# ── Sitemap loading ───────────────────────────────────────────────────────────

def _fetch_sitemap_urls(sitemap_url: str, depth: int = 0) -> list[str]:
    """Recursively load all product URLs from a sitemap or sitemap index."""
    if depth > 2:
        return []
    content = _fetch_direct(sitemap_url)
    if not content:
        return []

    sub_sitemaps = re.findall(r"<sitemap>\s*<loc>([^<]+)</loc>", content)
    if sub_sitemaps:
        product_subs = [s for s in sub_sitemaps if any(x in s.lower() for x in
                        ["product", "catalog", "acatalog", "p/", "sitemap-1-"])]
        to_fetch = product_subs if product_subs else sub_sitemaps
        to_fetch = [s for s in to_fetch if not any(x in s.lower() for x in
                    ["/news", "/blog", "/static-pages", "news.xml",
                     "categories.xml", "brands.xml", "static_pages"])]
        urls = []
        for sub in to_fetch:
            urls.extend(_fetch_sitemap_urls(html.unescape(sub.strip()), depth + 1))
            time.sleep(0.2)
        return urls

    all_locs = re.findall(r"<loc>([^<]+)</loc>", content)

    # Handle paginated sitemaps (e.g. sitemap_products.xml + sitemap_products_1.xml)
    if depth == 0 and len(all_locs) >= 50000:
        base = re.sub(r"\.xml.*$", "", sitemap_url)
        for n in range(1, 10):
            extra = _fetch_direct(f"{base}_{n}.xml")
            if not extra:
                break
            extra_locs = re.findall(r"<loc>([^<]+)</loc>", extra)
            if not extra_locs:
                break
            all_locs.extend(extra_locs)
            time.sleep(0.2)

    return all_locs


def _slug_to_title(url: str, domain: str) -> str:
    """Extract a human-readable title from a product URL slug."""
    path = url.replace(f"https://{domain}", "").replace(f"http://{domain}", "")
    path = path.lstrip("/")
    path = re.split(r"[?#]", path)[0]
    path = re.sub(r"\.(html?|php|aspx?)$", "", path)
    parts = [p for p in path.split("/") if p]

    slug = ""
    for part in reversed(parts):
        if re.match(r"^\d+$", part):                # pure numeric ID
            continue
        if len(part) <= 2:                           # single-char prefix (e.g. 'p')
            continue
        if re.match(r"^[a-f0-9]{8,}$", part):       # hash
            continue
        if re.match(r"^[a-z]{1,2}\d{3,}$", part):  # ID suffix e.g. p16927
            continue
        slug = part
        break

    if not slug and parts:
        slug = parts[-1]

    slug = re.sub(r"[-_]c\d{4,}$", "", slug)   # Oxatis IDs
    slug = re.sub(r"[-_]\d{4,}$", "", slug)     # trailing numeric IDs
    return slug.replace("-", " ").replace("_", " ").strip()


def build_url_map(comp_name: str, sitemap_url: str, domain: str) -> dict[str, str]:
    """Build {slug_title_lower: product_url} index from competitor sitemap."""
    print(f"  [{comp_name}] Loading sitemap...")
    all_urls = _fetch_sitemap_urls(sitemap_url)

    product_urls = []
    for u in all_urls:
        if not u.startswith("http"):
            continue
        path = u.replace(f"https://{domain}", "").replace(f"http://{domain}", "")
        if len([s for s in path.split("/") if s]) < 1:
            continue
        if any(x in path.lower() for x in ["/blog/", "/news/", "/help/", "/faq/",
                                             "/about", "/contact", "/sitemap", "/search"]):
            continue
        product_urls.append(u)

    url_map = {}
    for u in product_urls:
        title = _slug_to_title(u, domain)
        if len(title.split()) >= 2:
            url_map[title.lower()] = u

    print(f"  [{comp_name}] {len(url_map):,} products indexed")
    return url_map


# ── Matching & extraction ─────────────────────────────────────────────────────

def find_best_url(our_title: str, url_map: dict[str, str],
                  threshold: int = TITLE_MATCH_THRESHOLD) -> tuple[str | None, int]:
    best_score, best_url = 0, None
    our_lower = our_title.lower()
    for slug_title, url in url_map.items():
        score = max(
            fuzz.token_sort_ratio(our_lower, slug_title),
            fuzz.partial_ratio(our_lower, slug_title),
        )
        if score > best_score:
            best_score, best_url = score, url
    return (best_url, best_score) if best_score >= threshold else (None, best_score)


def _extract_price(content: str) -> float | None:
    """Extract first valid price from Jina markdown, skipping delivery thresholds."""
    price_pat = re.compile(r"[£\u00a3\ufffd]\s*(\d[\d,]*\.?\d{0,2})")
    for line in content.splitlines():
        if re.search(r"orders? over|free delivery|delivery from|shipping", line, re.IGNORECASE):
            continue
        for m in price_pat.finditer(line):
            val = _clean_price(m.group(1))
            if val and val > 2:
                return val
    return None


def _extract_page_title(content: str) -> str:
    for line in content.splitlines():
        m = re.match(r"^#{1,2}\s+(.+)$", line.strip())
        if m and len(m.group(1)) > 5:
            return m.group(1).strip()
    return ""


def _clean_price(raw: str) -> float | None:
    cleaned = re.sub(r"[^\d.]", "", raw)
    try:
        val = float(cleaned)
        return val if 0.50 < val < 50000 else None
    except ValueError:
        return None


def _title_score(a: str, b: str) -> int:
    if not a or not b:
        return 0
    return max(fuzz.token_sort_ratio(a.lower(), b.lower()),
               fuzz.partial_ratio(a.lower(), b.lower()))


# ── Main scrape loop ──────────────────────────────────────────────────────────

def run_scrape(work_list: list[dict], url_maps: dict) -> list[dict]:
    results = []
    total = len(work_list)

    for i, item in enumerate(work_list, 1):
        mpn       = item["mpn"]
        title     = item["title"]
        comp_name = item["competitor"]
        our_cost  = item.get("cost")
        domain    = COMPETITORS.get(comp_name, {}).get("domain", "")

        print(f"  [{i}/{total}] {comp_name:25s} {mpn:20s} {title[:35]}")

        url_map = url_maps.get(comp_name)
        if url_map is None:
            results.append(_row(mpn, title, item.get("brand", ""), comp_name, domain,
                                our_cost=our_cost, status="no_sitemap"))
            continue

        product_url, match_score = find_best_url(title, url_map)
        if not product_url:
            results.append(_row(mpn, title, item.get("brand", ""), comp_name, domain,
                                our_cost=our_cost, status=f"url_not_found_{match_score}pct"))
            continue

        content = _fetch(product_url)
        if not content:
            results.append(_row(mpn, title, item.get("brand", ""), comp_name, domain,
                                our_cost=our_cost, product_url=product_url, status="fetch_failed"))
            time.sleep(1)
            continue

        page_title = _extract_page_title(content)
        price      = _extract_price(content)
        page_score = _title_score(title, page_title)

        if not price:
            results.append(_row(mpn, title, item.get("brand", ""), comp_name, domain,
                                our_cost=our_cost, competitor_title=page_title,
                                title_match_score=page_score,
                                product_url=product_url, status="price_not_found"))
        elif our_cost and our_cost > 0 and price < our_cost * (1 - COST_FLOOR_TOLERANCE):
            print(f"    => £{price:.2f} BELOW COST FLOOR (our cost £{our_cost:.2f}) — rejected")
            results.append(_row(mpn, title, item.get("brand", ""), comp_name, domain,
                                our_cost=our_cost, competitor_title=page_title,
                                title_match_score=page_score, net_price=price,
                                product_url=product_url, status="below_cost_floor"))
        else:
            print(f"    => £{price:.2f} | \"{page_title[:50]}\" ({page_score}%)")
            results.append(_row(mpn, title, item.get("brand", ""), comp_name, domain,
                                our_cost=our_cost, competitor_title=page_title,
                                title_match_score=page_score, net_price=price,
                                shipping=0.0, total_price=price,
                                rank=1, product_url=product_url, status="matched"))
        time.sleep(1.2)

    return results


def _row(mpn, our_title, brand, competitor, domain, status="",
         our_cost=None, competitor_title="", title_match_score=0,
         net_price=None, shipping=None, total_price=None,
         rank=None, product_url=None):
    return {
        "mpn": mpn, "our_title": our_title, "brand": brand,
        "competitor": competitor, "competitor_domain": domain,
        "our_cost": our_cost,
        "competitor_title": competitor_title,
        "title_match_score": title_match_score,
        "net_price": net_price, "shipping": shipping, "total_price": total_price,
        "rank": rank, "product_url": product_url,
        "date_checked": datetime.now().strftime("%Y-%m-%d"),
        "source": "DirectScrape", "status": status,
    }


# ── Excel output ──────────────────────────────────────────────────────────────

def build_excel(results: list[dict], output_path: str):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    NAVY  = "1F3864"; WHITE = "FFFFFF"; GREEN = "E2EFDA"
    AMBER = "FFF2CC"; GREY  = "F2F2F2"; RED   = "FFE0E0"; ORANGE = "FFD9B3"

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Results ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Results"
    ws.sheet_view.showGridLines = False

    headers = ["MPN", "Brand", "Our Title", "Competitor", "Domain",
               "Our Cost", "Competitor Title", "Match %", "Net Price", "Shipping",
               "Total Price", "vs Cost", "Rank", "Product URL", "Date", "Status"]
    widths  = [16, 18, 40, 22, 26, 10, 40, 8, 11, 10, 11, 8, 6, 52, 12, 22]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    sev_bg = {"matched": GREEN, "no_sitemap": GREY, "fetch_failed": RED,
              "price_not_found": AMBER, "below_cost_floor": ORANGE}

    for rn, row in enumerate(results, 2):
        status    = row.get("status", "")
        bg        = next((v for k, v in sev_bg.items() if k in status), GREY)
        our_cost  = row.get("our_cost")
        net_price = row.get("net_price")
        vs_cost   = round(net_price / our_cost, 3) if net_price and our_cost and our_cost > 0 else None
        vals = [
            row["mpn"], row["brand"], row["our_title"], row["competitor"],
            row["competitor_domain"], our_cost, row["competitor_title"],
            row["title_match_score"] or "", net_price, row["shipping"],
            row["total_price"], vs_cost, row["rank"], row["product_url"],
            row["date_checked"], status,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=ci, value=v)
            c.font = Font(size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="top", wrap_text=(ci in (3, 7, 14)))
            if ci in (6, 9, 10, 11) and v is not None:
                c.number_format = "£#,##0.00"
            if ci == 12 and v is not None:
                c.number_format = "0.00x"
        ws.row_dimensions[rn].height = 18

    # ── Sheet 2: Price Matrix (MPN x Competitor) ──────────────────────────────
    ws2 = wb.create_sheet("Price Matrix")
    ws2.sheet_view.showGridLines = False

    matched      = [r for r in results if r["status"] == "matched"]
    comps_seen   = sorted({r["competitor"] for r in matched})
    mpns_seen    = list(dict.fromkeys(r["mpn"] for r in results))
    price_lookup = {(r["mpn"], r["competitor"]): r["total_price"] for r in matched}
    cost_lookup  = {r["mpn"]: r.get("our_cost") for r in results}

    hdrs = ["MPN", "Brand", "Our Title", "Our Cost"] + comps_seen
    for i, h in enumerate(hdrs, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 10
    for i in range(5, len(hdrs) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.row_dimensions[1].height = 22

    mpn_info = {r["mpn"]: (r["brand"], r["our_title"]) for r in results}
    for rn, mpn in enumerate(mpns_seen, 2):
        brand, title = mpn_info.get(mpn, ("", ""))
        our_cost = cost_lookup.get(mpn)
        ws2.cell(row=rn, column=1, value=mpn).font = Font(size=10)
        ws2.cell(row=rn, column=2, value=brand).font = Font(size=10)
        ws2.cell(row=rn, column=3, value=title).font = Font(size=10)
        c = ws2.cell(row=rn, column=4, value=our_cost)
        c.font = Font(size=10)
        if our_cost:
            c.number_format = "£#,##0.00"
        for ci, comp in enumerate(comps_seen, 5):
            price = price_lookup.get((mpn, comp))
            c = ws2.cell(row=rn, column=ci, value=price)
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal="center")
            if price is not None:
                c.number_format = "£#,##0.00"
                c.fill = PatternFill("solid", fgColor=GREEN)
            else:
                c.value = "—"
                c.fill = PatternFill("solid", fgColor=GREY)
        ws2.row_dimensions[rn].height = 18

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 12

    for i, h in enumerate(["Competitor", "Matched", "Total Jobs"], 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22

    total_by_comp   = Counter(r["competitor"] for r in results)
    matched_by_comp = Counter(r["competitor"] for r in results if r["status"] == "matched")

    for rn, (comp, total) in enumerate(
            sorted(total_by_comp.items(), key=lambda x: -matched_by_comp.get(x[0], 0)), 2):
        n_matched = matched_by_comp.get(comp, 0)
        bg = GREEN if n_matched > 0 else GREY
        for ci, v in enumerate([comp, n_matched, total], 1):
            c = ws3.cell(row=rn, column=ci, value=v)
            c.font = Font(size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
        ws3.row_dimensions[rn].height = 18

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ── Data loaders ──────────────────────────────────────────────────────────────

def load_catalog(path: str) -> dict[str, dict]:
    """Load MPN -> {title, brand, revenue, cost} from Price Vs Cost export."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    if "long_description" in idx:
        title_col = idx["long_description"]
        mpn_col   = idx["product"]
        brand_col = idx.get("Company", idx.get("supplier"))
        rev_col   = idx.get("GraphBars")
        cost_col  = idx.get("WeightedAverageCost", idx.get("Cost"))
    else:
        title_col, mpn_col, brand_col, rev_col, cost_col = 1, 2, 3, None, None

    products = {}
    for row in ws.iter_rows(values_only=True, min_row=2):
        mpn   = str(row[mpn_col]).strip() if row[mpn_col] else None
        title = str(row[title_col]).strip() if row[title_col] else None
        brand = str(row[brand_col]).strip() if brand_col is not None and row[brand_col] else ""
        rev   = float(row[rev_col] or 0) if rev_col is not None else 0
        cost  = float(row[cost_col] or 0) if cost_col is not None and row[cost_col] else None
        if not mpn or not title or mpn in ("None", "product"):
            continue
        if mpn not in products or rev > products[mpn]["revenue"]:
            products[mpn] = {"title": title, "brand": brand, "revenue": rev, "cost": cost}
    return products


def load_matches(path: str) -> dict[str, set]:
    """Load MPN -> {competitor names} from Competitor Matching export."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    mpn_comps: dict = {}
    for row in ws.iter_rows(values_only=True, min_row=2):
        _, comp, _, _, _, mpn = row
        if mpn and comp:
            mpn_comps.setdefault(str(mpn), set()).add(str(comp))
    return mpn_comps


def build_work_list(catalog: dict, mpn_comps: dict, top_n: int) -> list[dict]:
    """Build scrape jobs sorted by revenue descending, capped at top_n (0 = all)."""
    work = []
    for mpn, info in catalog.items():
        for comp in mpn_comps.get(mpn, set()):
            if comp in COMPETITORS:
                work.append({
                    "mpn":        mpn,
                    "title":      info["title"],
                    "brand":      info.get("brand", ""),
                    "revenue":    info.get("revenue", 0),
                    "cost":       info.get("cost"),
                    "competitor": comp,
                })
    work.sort(key=lambda x: x["revenue"], reverse=True)
    return work[:top_n] if top_n > 0 else work


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    top_n = int(sys.argv[1]) if len(sys.argv) > 1 else TOP_N_PRODUCTS

    print(f"PriceAPI Clone — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"Loading catalog and matches...")

    catalog   = load_catalog(CATALOG_PATH)
    matches   = load_matches(MATCHES_PATH)
    work_list = build_work_list(catalog, matches, top_n)

    comp_counts = Counter(w["competitor"] for w in work_list)
    print(f"\n{len(work_list)} jobs across {len(comp_counts)} competitors:")
    for comp, n in comp_counts.most_common():
        has_sitemap = "sitemap" if COMPETITORS.get(comp, {}).get("sitemap") else "no sitemap"
        print(f"  {n:4d}  {comp:30s} ({has_sitemap})")

    url_maps: dict = {}
    comps_needed = {w["competitor"] for w in work_list}
    print("\nLoading sitemaps...")
    for comp_name in comps_needed:
        cfg = COMPETITORS.get(comp_name, {})
        sitemap_url = cfg.get("sitemap")
        domain = cfg.get("domain", "")
        if sitemap_url:
            url_maps[comp_name] = build_url_map(comp_name, sitemap_url, domain)
        else:
            url_maps[comp_name] = None

    print(f"\nScraping {len(work_list)} jobs...")
    results = run_scrape(work_list, url_maps)

    matched = sum(1 for r in results if r["status"] == "matched")
    print(f"\n{matched}/{len(results)} prices retrieved")

    os.makedirs(REPORTS_DIR, exist_ok=True)
    fname = f"CompetitorPrices-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    build_excel(results, os.path.join(REPORTS_DIR, fname))
    print("Done. Report saved to WORKAI/reports/ — Google Drive will sync automatically.")
